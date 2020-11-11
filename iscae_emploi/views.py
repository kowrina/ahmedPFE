from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render, redirect


from datetime import datetime, timedelta
from dateutil import rrule
import mimetypes
import pandas as pd

from .models import *

from .forms import *
from django import forms

from .filters import *
# Create your views here.

#Programme lineaire librairie

from ortools.linear_solver import pywraplp
from openpyxl import Workbook #pip install openxl
from openpyxl.styles import Font,Color,colors,PatternFill,Border,Side,Alignment,fills
import  numpy as np



@login_required(login_url='/')
def handler404(request,exception = None, templates_name = '404.html'):
    return render(request, '404.html', status=404)

@login_required(login_url='/')
def handler500(request,exception = None, templates_name = '500.html'):
    return render(request, '500.html', status=500)




@login_required(login_url='/')
def index(request):

    profs = Professeur.objects.all()
    deps = Departement.objects.all()
    nombre_profs = profs.count()
    dep1 = deps[0]
    prof_dep1 = Professeur.objects.filter(dep = dep1)
    nombre_profs_dep1 = prof_dep1.count()
    dep1_pourcentage = int(nombre_profs_dep1 * 100 /nombre_profs)

    dep2 = deps[1]
    prof_dep2 = Professeur.objects.filter(dep = dep2)
    nombre_profs_dep2 = prof_dep2.count()
    dep2_pourcentage = int(nombre_profs_dep2 * 100 /nombre_profs)

    dep3 = deps[2]
    prof_dep3 = Professeur.objects.filter(dep = dep3)
    nombre_profs_dep3 = prof_dep3.count()
    dep3_pourcentage = int(nombre_profs_dep3 * 100 /nombre_profs)

    dep4 = deps[3]
    prof_dep4 = Professeur.objects.filter(dep = dep4)
    nombre_profs_dep4 = prof_dep4.count()
    dep4_pourcentage = int(nombre_profs_dep4 * 100 /nombre_profs)

    request.user

    return render(request,'index.html',locals())



@login_required(login_url='/')
def config(request):
    return render(request,'config.html',{})


@login_required(login_url='/')
def departement(request):
    departements = Departement.objects.all()

    pagination = Paginator(departements,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'Departements',
        'departements': page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'departments.html',context)


@login_required(login_url='/')
def prof(request,pk):
    profs = Professeur.objects.filter(dep__id=pk)
    dep = Departement.objects.get(id=pk)

    pagination = Paginator(profs,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'Professeur',
        'dep': dep ,
        'profs': page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'professeur.html',context)




@login_required(login_url='/')
def dep_matiers(request,pk):
    matier = Matier.objects.filter(dep__id=pk)
    dep = Departement.objects.get(id=pk)

    pagination = Paginator(matier,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)
    context = {
        'title': 'matier',
        'matier':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'dep' : dep,
    }
    return render(request,'dep_matier.html',context)

@login_required(login_url='/')
def classe_matiere(request,pk):
    classe = Classe.objects.get(id=pk)
    matier = classe.matier_etu.all()

    pagination = Paginator(matier,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'matier',
        'matier':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'classe' : classe,

    }
    return render(request,'classe_matier.html',context)


@login_required(login_url='/')
def classe_groupe(request,pk):
    groupe = Groupe.objects.filter(classe__id=pk)
    classe = Classe.objects.get(id=pk)

    pagination = Paginator(groupe,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'groupes',
        'groupe':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'classe' : classe,
    }
    return render(request,'classe_groupe.html',context)


@login_required(login_url='/')
def filier_classes(request,pk):
    filier_classes = Classe.objects.filter(filier__id=pk)
    filier = Filier.objects.get(id=pk)

    pagination = Paginator(filier_classes,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'dp_classes',
        'filier_classes':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'filier' : filier,
    }
    return render(request,'filier_classes.html',context)



@login_required(login_url='/')
def cours(request):
    cour = Cours.objects.all()

    pagination = Paginator(cour,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'cour',
        'cour':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'cours.html',context)






@login_required(login_url='/')
def groupe_cours(request,pk):
    groupe_cours = Cours.objects.filter(groupe__id=pk)
    groupe = Groupe.objects.get(id=pk)

    pagination = Paginator(groupe_cours,per_page=4)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'groupe cours',
        'groupe_cours':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'groupe' : groupe ,
    }
    return render(request,'groupe_cours.html',context)


@login_required(login_url='/')
def groupe_indispo(request,pk):
    g_indsipo =  IndispoGroupe.objects.filter(groupe__id=pk)
    groupe= Groupe.objects.get(id=pk)

    pagination = Paginator(g_indsipo,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'prof indispo',
        'g_indispo':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'groupe':groupe,
    }
    return render(request,'groupe_indispo.html',context)


@login_required(login_url='/')
def groupe_indispo_creneau(request,pk):
    g_indsipo =  IndispoGroupe.objects.get(id=pk)
    creneau= g_indsipo.créneaux.all()

    pagination = Paginator(creneau,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'prof indispo Creneau',
        'creneau':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),

    }
    return render(request,'groupe_indispo_creneau.html',context)



@login_required(login_url='/')
def prof_cours(request,pk):
    p_cours =  Cours.objects.filter(prof__id=pk)

    pagination = Paginator(p_cours,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'p_cours',
        'p_cours':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'prof_cours.html',context)


@login_required(login_url='/')
def prof_indispo(request,pk):
    p_indsipo =  IndispoProf.objects.filter(prof__id=pk)
    prof= Professeur.objects.get(id=pk)

    pagination = Paginator(p_indsipo,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'prof indispo',
        'p_indispo':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'prof':prof,
    }
    return render(request,'prof_indispo.html',context)


@login_required(login_url='/')
def prof_indispo_creneau(request,pk):
    p_indsipo =  IndispoProf.objects.get(id=pk)
    creneau= p_indsipo.créneaux.all()

    pagination = Paginator(creneau,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'prof indispo Creneau',
        'creneau':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),

    }
    return render(request,'prof_indispo_creneau.html',context)

@login_required(login_url='/')
def prof_preferance(request,pk):
    p_preferance =  Preferance_prof.objects.filter(prof__id=pk)
    prof= Professeur.objects.get(id=pk)

    pagination = Paginator(p_preferance,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'prof preferance',
        'p_preferance':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'prof':prof,
    }
    return render(request,'prof_preferance.html',context)


@login_required(login_url='/')
def prof_preferance_creneau(request,pk):
    p_preferance =  Preferance_prof.objects.get(id=pk)
    creneau= p_preferance.créneaux.all()

    pagination = Paginator(creneau,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'prof preferance Creneau',
        'creneau':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),

    }
    return render(request,'prof_indispo_creneau.html',context)


@login_required(login_url='/')
def salle(request):
    salles = Salle.objects.all()

    pagination = Paginator(salles,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'salles',
        'salles':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'salle.html',context)



@login_required(login_url='/')
def filiere(request,pk):
    filieres = Filier.objects.filter(dep__id = pk)
    dep = Departement.objects.get(id = pk)

    pagination = Paginator(filieres,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'filieres',
        'filieres': page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'dep':dep,
    }
    return render(request,'filieres.html',context)




@login_required(login_url='/')
def niveau(request):
    niveaux = Niveau.objects.all()

    pagination = Paginator(niveaux,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'niveaux',
        'niveaux': page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'niveaux.html',context)



@login_required(login_url='/')
def jours(request):
    jour = Jour.objects.all()

    pagination = Paginator(jour,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'jour',
        'jour': page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'jours.html',context)

@login_required(login_url='/')
def creneaux(request):
    creneau = Creneau.objects.all()

    pagination = Paginator(creneau,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'creneau',
        'creneau':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'creneaux.html',context)


@login_required(login_url='/')
def j_creneaux(request,pk):
    detail =  DetailCalandrier.objects.get(jour__id=pk)
    creneau = detail.créneaux.all()
    jour = Jour.objects.get(id=pk)

    pagination = Paginator(creneau,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)
    context = {
        'title': 'creneau',
        'creneau':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'jour':jour,
    }
    return render(request,'j_creneaux.html',context)




@login_required(login_url='/')
def calandriers(request):
    cal = CalandrierStandard.objects.all()

    pagination = Paginator(cal,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'calandrier',
        'cal': page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
    }
    return render(request,'calandriers_Stdr.html',context)

@login_required(login_url='/')
def detail_c(request,pk):
    detail = DetailCalandrier.objects.filter(calandrier__id=pk)
    cal=CalandrierStandard.objects.get(id=pk)

    pagination = Paginator(detail,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'detail',
        'detail': page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'cal':cal,
    }
    return render(request,'detail_c.html',context)


@login_required(login_url='/')
def c_exception(request,pk):
    exception = ExceptionCalandrier.objects.filter(calandrier__id=pk)
    cal = CalandrierStandard.objects.get(id=pk)

    pagination = Paginator(exception,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)

    context = {
        'title': 'exception' ,
        'exception': page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'cal':cal,
    }
    return render(request,'c_exception.html',context)




#*******************************  Add Form *********************************




#ajout de departement
@login_required(login_url='/')
def dep_ajout(request):

    form=DepartementForm()
    if request.method == 'POST':
        form = DepartementForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/dp/')
    else:
        form = DepartementForm()

    context = {
        'title': 'Ajout Dep',
        'form': form,
    }
    return render(request,'dep_ajout.html',context)

#edit de departement
@login_required(login_url='/')
def dep_edit(request,pk):

    dep=Departement.objects.get(id=pk)
    form=DepartementForm(instance=dep)

    if request.method == 'POST':
        form = DepartementForm(request.POST,instance=dep)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/dp/')
    else:
        form = DepartementForm(instance=dep)

    context = {
        'title': 'Edit Dep',
        'form': form,
        'dep':dep,
    }
    return render(request,'dep_edit.html',context)

@login_required(login_url='/')
def dep_delete(request,pk):
    dep = Departement.objects.get(id=pk)

    if request.method == 'POST':
        dep.delete()
        return redirect('/iscae_emploi/dp/')

    context = {
        'title': 'Delete Dep',
        'item': dep,
    }
    return render(request,'dep_delete.html',context)

#ajout de niveau
@login_required(login_url='/')
def niveau_ajout(request):

    form=NiveauForm()
    if request.method == 'POST':
        form = NiveauForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/niveaux/')
    else:
        form = NiveauForm()

    context = {
        'title': 'Ajout niveau',
        'form': form,
    }
    return render(request,'niveau_ajout.html',context)

#edit de niveau
@login_required(login_url='/')
def niveau_edit(request,pk):

    niveau=Niveau.objects.get(id=pk)
    form=NiveauForm(instance=niveau)

    if request.method == 'POST':
        form = NiveauForm(request.POST,instance=niveau)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/niveaux/')
    else:
        form = NiveauForm(instance=niveau)

    context = {
        'title': 'Edit niveau',
        'form': form,
        'niveau':niveau,
    }
    return render(request,'niveau_edit.html',context)

#niveau delete
@login_required(login_url='/')
def niveau_delete(request,pk):
    niveau = Niveau.objects.get(id=pk)

    if request.method == 'POST':
        niveau.delete()
        return redirect('/iscae_emploi/niveaux/')

    context = {
        'title': 'Delete niveau',
        'item': niveau,
    }
    return render(request,'niveau_delete.html',context)



## ajout de sall
@login_required(login_url='/')
def sall_ajout(request):

    form=SalleForm()
    if request.method == 'POST':
        form = SalleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/salles/')
    else:
        form = SalleForm()

    context = {
        'title': 'Ajout sall',
        'form': form,
    }
    return render(request,'sall_ajout.html',context)



#edit de sall
@login_required(login_url='/')
def sall_edit(request,pk):

    sall=Salle.objects.get(id=pk)
    form=SalleForm(instance=sall)

    if request.method == 'POST':
        form = SalleForm(request.POST,instance=sall)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/salles/')
    else:
        form = SalleForm(instance=sall)

    context = {
        'title': 'Edit sall',
        'form': form,
        'sall':sall,
    }
    return render(request,'sall_edit.html',context)


#delete de sall
@login_required(login_url='/')
def sall_delete(request,pk):
    sall = Salle.objects.get(id=pk)

    if request.method == 'POST':
        sall.delete()
        return redirect('/iscae_emploi/salles/')

    context = {
        'title': 'Delete Dep',
        'item': sall,
    }
    return render(request,'sall_delete.html',context)


## ajout de cal
@login_required(login_url='/')
def calandrier_ajout(request):
    form=CalandrierForm()
    if request.method == 'POST':
        form = CalandrierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/calandriers/')
    else:
        form = CalandrierForm()

    context = {
        'title': 'Ajout Calandrier',
        'form': form,
    }
    return render(request,'cal_ajout.html',context)



#edit de calandrier
@login_required(login_url='/')
def calandrier_edit(request,pk):

    cal=CalandrierStandard.objects.get(id=pk)
    form=CalandrierForm(instance=cal)

    if request.method == 'POST':
        form = CalandrierForm(request.POST,instance=cal)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/calandriers/')
    else:
        form = CalandrierForm(instance=cal)

    context = {
        'title': 'Edit calandrier',
        'form': form,
        'cal':cal,
    }
    return render(request,'cal_edit.html',context)


#delete de calandrier
@login_required(login_url='/')
def calandrier_delete(request,pk):
    cal = CalandrierStandard.objects.get(id=pk)

    if request.method == 'POST':
        cal.delete()
        return redirect('/iscae_emploi/calandriers/')

    context = {
        'title': 'Delete calandrier',
        'item': cal,
    }
    return render(request,'cal_delete.html',context)


#ajout detail

@login_required(login_url='/')
def detail_c_ajout(request,pk):

    cal = CalandrierStandard.objects.get(id=pk)


    form=DetailcalandrierForm()
    if request.method == 'POST':
        form = DetailcalandrierForm(request.POST)

        if form.is_valid():

            detail = form.save(commit=False)
            detail.calandrier=cal
            detail.save()
        return redirect('/iscae_emploi/detail_c/%i' %pk)

    context = {
        'title': 'Ajout Detail',
        'cal': cal,
        'form':form
    }
    return render(request,'detail_c_ajout.html',context)


#edit de detail
@login_required(login_url='/')
def detail_c_edit(request,pk):

    detail=DetailCalandrier.objects.get(id=pk)
    cal = detail.calandrier
    form=DetailcalandrierForm(instance=detail)

    if request.method == 'POST':
        form = DetailcalandrierForm(request.POST,instance=detail)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/detail_c/%i' %cal.id)
    else:
        form = DetailcalandrierForm(instance=detail)

    context = {
        'title': 'Edit detail',
        'form': form,
        'detail':detail,
    }
    return render(request,'detail_c_edit.html',context)


#delete de detail
@login_required(login_url='/')
def detail_c_delete(request,pk):
    detail = DetailCalandrier.objects.get(id=pk)
    cal=detail.calandrier

    if request.method == 'POST':
        detail.delete()
        return redirect('/iscae_emploi/detail_c/%i' %cal.id)

    context = {
        'title': 'Delete Detail',
        'item': detail,
    }
    return render(request,'detail_c_delete.html',context)


## ajout de creneau
@login_required(login_url='/')
def creneau_ajout(request):

    form=CreneauForm()

    if request.method == 'POST':
        form = CreneauForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/creneaux/')
    else:
        form = CreneauForm()

    context = {
        'title': 'Ajout Creneau',
        'form': form,
    }
    return render(request,'creneau_ajout.html',context)



#edit de creneau
@login_required(login_url='/')
def creneau_edit(request,pk):

    creneau=Creneau.objects.get(id=pk)
    form=CreneauForm(instance=creneau)

    if request.method == 'POST':
        form = CreneauForm(request.POST,instance=creneau)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/creneaux/')
    else:
        form = CreneauForm(instance=creneau)

    context = {
        'title': 'Edit creneau',
        'form': form,
        'creneau':creneau,
    }
    return render(request,'creneau_edit.html',context)


#delete de Creneau
@login_required(login_url='/')
def creneau_delete(request,pk):
    creneau = Creneau.objects.get(id=pk)

    if request.method == 'POST':
        creneau.delete()
        return redirect('/iscae_emploi/creneaux/')

    context = {
        'title': 'Delete Creneau',
        'item': creneau,
    }
    return render(request,'creneau_delete.html',context)

#ajout matiere

@login_required(login_url='/')
def matier_ajout(request,pk):

    dep = Departement.objects.get(id=pk)


    form=MatierForm()
    if request.method == 'POST':
        form = MatierForm(request.POST)
        if form.is_valid():

            matier = form.save(commit=False)
            matier.dep=dep
            matier.save()
        return redirect('/iscae_emploi/dep_matiers/%i' %pk)


    context = {
        'title': 'Ajout matiere',
        'dep': dep,
        'form':form
    }
    return render(request,'matier_ajout.html',context)


#edit de matiere
@login_required(login_url='/')
def matier_edit(request,pk):

    matier=Matier.objects.get(id=pk)
    dep = matier.dep
    form=MatierForm(instance=matier)

    if request.method == 'POST':
        form = MatierForm(request.POST,instance=matier)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/dep_matiers/%i' %dep.id)
    else:
        form = MatierForm(instance=matier)

    context = {
        'title': 'Edit matiere',
        'form': form,
        'matier':matier,
    }
    return render(request,'matier_edit.html',context)


#delete de matiere
@login_required(login_url='/')
def matier_delete(request,pk):
    matier = Matier.objects.get(id=pk)
    dep = matier.dep

    if request.method == 'POST':
        matier.delete()
        return redirect('/iscae_emploi/dep_matiers/%i' %dep.id)

    context = {
        'title': 'Delete Matiere',
        'item': matier,
    }
    return render(request,'matier_delete.html',context)



#ajout prof

@login_required(login_url='/')
def prof_ajout(request,pk):

    dep = Departement.objects.get(id=pk)


    form=ProfesseurForm()
    if request.method == 'POST':
        form = ProfesseurForm(request.POST)
        if form.is_valid():

            prof = form.save(commit=False)
            prof.dep=dep
            prof.save()
        return redirect('/iscae_emploi/prof/%i'%pk)


    context = {
        'title': 'Ajout professeur',
        'dep': dep,
        'form':form,
    }
    return render(request,'prof_ajout.html',context)


#edit de prof
@login_required(login_url='/')
def prof_edit(request,pk):

    prof=Professeur.objects.get(id=pk)
    dep=prof.dep
    form=ProfesseurForm(instance=prof)

    if request.method == 'POST':
        form = ProfesseurForm(request.POST,instance=prof)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/prof/%i' %dep.id)
    else:
        form = ProfesseurForm(instance=prof)

    context = {
        'title': 'Edit professeur',
        'form': form,
        'prof':prof,
    }
    return render(request,'prof_edit.html',context)


#delete de prof
@login_required(login_url='/')
def prof_delete(request,pk):
    prof = Professeur.objects.get(id=pk)
    dep=prof.dep

    if request.method == 'POST':
        prof.delete()
        return redirect('/iscae_emploi/prof/%i' %dep.id)

    context = {
        'title': 'supprimer professeur',
        'item': prof,
    }
    return render(request,'prof_delete.html',context)


#ajout prof_indispo_

@login_required(login_url='/')
def prof_indispo_ajout(request,pk):
    prof = Professeur.objects.get(id=pk)

    form=IndispoprofForm()
    if request.method == 'POST':
        form = IndispoprofForm(request.POST)
        if form.is_valid():

            indispo = form.save(commit=False)
            indispo.prof=prof
            indispo.date= request.POST['date']
            indispo.save()
            form.save_m2m()
        return redirect('/iscae_emploi/prof_indispo/%i' %pk)


    context = {
        'title': 'Ajout indispo',
        'prof': prof,
        'form':form
    }
    return render(request,'prof_indispo_ajout.html',context)


#edit de prof_indispo_
@login_required(login_url='/')
def prof_indispo_edit(request,pk):

    indispo=IndispoProf.objects.get(id=pk)
    prof = indispo.prof
    form=IndispoprofForm(instance=indispo)

    if request.method == 'POST':
        form = IndispoprofForm(request.POST,instance=indispo)
        if form.is_valid():
            print("------------------------->")
            print(form.fields)
            print("------------------------->")
            indisp = form.save(commit=False)
            indisp.date= request.POST['date']
            indisp.save()
            form.save_m2m()
            return redirect('/iscae_emploi/prof_indispo/%i' %prof.id)
    else:
        form = IndispoprofForm(instance=indispo)

    context = {
        'title': 'Edit indispo',
        'form': form,
        'indispo':indispo,
        'prof':prof
    }
    return render(request,'prof_indispo_edit.html',context)


#delete de indispo prof
@login_required(login_url='/')
def prof_indispo_delete(request,pk):
    indispo = IndispoProf.objects.get(id=pk)
    prof = indispo.prof

    if request.method == 'POST':
        indispo.delete()
        return redirect('/iscae_emploi/prof_indispo/%i' %prof.id)

    context = {
        'title': 'Delete indispo',
        'item': indispo,
        'prof':prof,
    }
    return render(request,'prof_indispo_delete.html',context)


#ajout prof_preferance_

@login_required(login_url='/')
def prof_preferance_ajout(request,pk):
    prof = Professeur.objects.get(id=pk)

    form=PreferanceprofForm()
    if request.method == 'POST':
        form = PreferanceprofForm(request.POST)
        if form.is_valid():

            preferance = form.save(commit=False)
            preferance.prof=prof
            preferance.save()
        return redirect('/iscae_emploi/prof_preferance/%i' %pk)


    context = {
        'title': 'Ajout preferance',
        'prof': prof,
        'form':form
    }
    return render(request,'prof_preferance_ajout.html',context)


#edit de prof_preferance
@login_required(login_url='/')
def prof_preferance_edit(request,pk):

    preferance=IndispoProf.objects.get(id=pk)
    prof = preferance.prof
    form=PreferanceprofForm(instance=preferance)

    if request.method == 'POST':
        form = PreferanceprofForm(request.POST,instance=preferance)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/prof_preferance/%i' %prof.id)
    else:
        form = PreferanceprofForm(instance=preferance)

    context = {
        'title': 'Edit preferance',
        'form': form,
        'preferance':preferance,
        'prof':prof
    }
    return render(request,'prof_preferance_edit.html',context)


#delete de preferance prof
@login_required(login_url='/')
def prof_preferance_delete(request,pk):
    preferance = Preferance_prof.objects.get(id=pk)
    prof = preferance.prof

    if request.method == 'POST':
        preferance.delete()
        return redirect('/iscae_emploi/prof_preferance/%i' %prof.id)

    context = {
        'title': 'Delete indispo',
        'item': preferance,
        'prof':prof,
    }
    return render(request,'prof_preferance_delete.html',context)

#ajout filiere

@login_required(login_url='/')
def filier_ajout(request,pk):

    dep = Departement.objects.get(id=pk)

    form=FilierForm()
    if request.method == 'POST':
        form = FilierForm(request.POST)
        if form.is_valid():

            filiere = form.save(commit=False)
            filiere.dep=dep
            filiere.save()
        return redirect('/iscae_emploi/filiere/%i'%pk)

    context = {
        'title': 'Ajout filiere',
        'dep': dep,
        'form':form,
    }
    return render(request,'filier_ajout.html',context)


#edit de filiere
@login_required(login_url='/')
def filier_edit(request,pk):

    filier=Filier.objects.get(id=pk)
    dep=filier.dep
    form=FilierForm(instance=filier)

    if request.method == 'POST':
        form = FilierForm(request.POST,instance=filier)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/filiere/%i' %dep.id)
    else:
        form = FilierForm(instance=filier)

    context = {
        'title': 'Edit filiere',
        'form': form,
        'filier':filier,
    }
    return render(request,'filier_edit.html',context)


#delete de filiere
@login_required(login_url='/')
def filier_delete(request,pk):
    filier=Filier.objects.get(id=pk)
    dep=filier.dep

    if request.method == 'POST':
        filier.delete()
        return redirect('/iscae_emploi/filiere/%i' %dep.id)

    context = {
        'title': 'supprimer filiere',
        'item': filier,
    }
    return render(request,'filier_delete.html',context)



#ajout classe

@login_required(login_url='/')
def classe_ajout(request,pk):
    filier = Filier.objects.get(id=pk)

    form=ClasseForm()
    if request.method == 'POST':
        form = ClasseForm(request.POST)
        if form.is_valid():

            classe = form.save(commit=False)
            classe.filier=filier
            classe.save()
        return redirect('/iscae_emploi/filier_classes/%i'%pk)

    context = {
        'title': 'Ajout filiere',
        'filier': filier,
        'form':form,
    }
    return render(request,'classe_ajout.html',context)


#edit de classe
@login_required(login_url='/')
def classe_edit(request,pk):

    classe=Classe.objects.get(id=pk)
    filier=classe.filier
    form=ClasseForm(instance=classe)

    if request.method == 'POST':
        form = ClasseForm(request.POST,instance=classe)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/filier_classes/%i' %filier.id)
    else:
        form = ClasseForm(instance=filier)

    context = {
        'title': 'Edit filiere',
        'form': form,
        'classe':classe,
    }
    return render(request,'classe_edit.html',context)


#delete de classe
@login_required(login_url='/')
def classe_delete(request,pk):
    classe=Classe.objects.get(id=pk)
    filier=classe.filier

    if request.method == 'POST':
        classe.delete()
        return redirect('/iscae_emploi/filier_classes/%i' %filier.id)

    context = {
        'title': 'supprimer un classe',
        'item': classe,
    }
    return render(request,'classe_delete.html',context)




## ajout de jour
@login_required(login_url='/')
def joure_ajout(request):

    form=JourForm()
    if request.method == 'POST':
        form = JourForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/jours/')
    else:
        form = JourForm()

    context = {
        'title': 'Ajout jour',
        'form': form,
    }
    return render(request,'joure_ajout.html',context)



#edit de jour
@login_required(login_url='/')
def joure_edit(request,pk):

    jour=Jour.objects.get(id=pk)
    form=JourForm(instance=jour)

    if request.method == 'POST':
        form = JourForm(request.POST,instance=jour)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/jours/')
    else:
        form = JourForm(instance=jour)

    context = {
        'title': 'Edit sall',
        'form': form,
        'jour':jour,
    }
    return render(request,'joure_edit.html',context)


#delete de jour
@login_required(login_url='/')
def joure_delete(request,pk):

    jour = Jour.objects.get(id=pk)

    if request.method == 'POST':
        jour.delete()
        return redirect('/iscae_emploi/jours/')

    context = {
        'title': 'Delete Dep',
        'item': jour,
    }
    return render(request,'joure_delete.html',context)


#ajout cours

@login_required(login_url='/')
def cours_ajout(request,pk):
    groupe = Groupe.objects.get(id=pk)

    form=CoursForm()
    if request.method == 'POST':
        form = CoursForm(request.POST)
        if form.is_valid():

            cours = form.save(commit=False)
            cours.groupe=groupe
            cours.save()
        return redirect('/iscae_emploi/groupe_cours/%i'%pk)

    context = {
        'title': 'Ajout cours',
        'groupe': groupe,
        'form':form,
    }
    return render(request,'cours_ajout.html',context)


#edit de cours
@login_required(login_url='/')
def cours_edit(request,pk):

    cours=Cours.objects.get(id=pk)

    form=CoursForm(instance=cours)

    if request.method == 'POST':
        form = CoursForm(request.POST,instance=cours)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/Cours')
    else:
        form = CoursForm(instance=cours)

    context = {
        'title': 'Edit cours',
        'form': form,
        'cours':cours,
    }
    return render(request,'cours_edit.html',context)


#delete de cours
@login_required(login_url='/')
def cours_delete(request,pk):
    cours=Cours.objects.get(id=pk)

    if request.method == 'POST':
        cours.delete()
        return redirect('/iscae_emploi/Cours')

    context = {
        'title': 'supprimer un cours',
        'item': cours,
    }
    return render(request,'cours_delete.html',context)


#ajout groupe

@login_required(login_url='/')
def groupe_ajout(request,pk):

    classe = Classe.objects.get(id=pk)

    form=GroupeForm()
    if request.method == 'POST':
        form = GroupeForm(request.POST)
        if form.is_valid():

            groupe = form.save(commit=False)
            groupe.classe=classe
            groupe.save()
        return redirect('/iscae_emploi/classe_groupe/%i'%pk)

    context = {
        'title': 'Ajout groupe',
        'classe': classe,
        'form': form,
    }
    return render(request,'groupe_ajout.html',context)


#edit de filiere
@login_required(login_url='/')
def groupe_edit(request,pk):

    groupe=Groupe.objects.get(id=pk)
    classe=groupe.classe
    form=GroupeForm(instance=groupe)

    if request.method == 'POST':
        form = GroupeForm(request.POST,instance=groupe)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/classe_groupe/%i' %classe.id)
    else:
        form = GroupeForm(instance=groupe)

    context = {
        'title': 'Edit Groupe',
        'form': form,
        'groupe':groupe,
    }
    return render(request,'groupe_edit.html',context)


#delete de groupe
@login_required(login_url='/')
def groupe_delete(request,pk):
    groupe=Groupe.objects.get(id=pk)
    classe=groupe.classe

    if request.method == 'POST':
        groupe.delete()
        return redirect('/iscae_emploi/classe_groupe/%i' %classe.id)

    context = {
        'title': 'supprimer filiere',
        'item': groupe,
        'classe':classe,
    }
    return render(request,'groupe_delete.html',context)


#ajout groupe_indispo_

@login_required(login_url='/')
def groupe_indispo_ajout(request,pk):
    groupe = Groupe.objects.get(id=pk)

    form=IndispogroupeForm()
    if request.method == 'POST':
        form = IndispogroupeForm(request.POST)
        if form.is_valid():

            indispo = form.save(commit=False)
            indispo.groupe=groupe
            indispo.date= request.POST['date']
            indispo.save()
            form.save_m2m()
        return redirect('/iscae_emploi/groupe_indispo/%i' %pk)


    context = {
        'title': 'Ajout indispo',
        'groupe': groupe,
        'form':form
    }
    return render(request,'groupe_indispo_ajout.html',context)


#edit de groupe_indispo_
@login_required(login_url='/')
def groupe_indispo_edit(request,pk):

    indispo=IndispoGroupe.objects.get(id=pk)
    groupe = indispo.groupe
    form=IndispogroupeForm(instance=indispo)

    if request.method == 'POST':
        form = IndispogroupeForm(request.POST,instance=indispo)
        if form.is_valid():
            indisp = form.save(commit=False)
            indisp.date= request.POST['date']
            indisp.save()
            form.save_m2m()
            return redirect('/iscae_emploi/groupe_indispo/%i' %groupe.id)
    else:
        form = IndispogroupeForm(instance=indispo)

    context = {
        'title': 'Edit indispo',
        'form': form,
        'indispo':indispo,
        'groupe':groupe
    }
    return render(request,'groupe_indispo_edit.html',context)


#delete de groupe_indispo_
@login_required(login_url='/')
def groupe_indispo_delete(request,pk):
    indispo = IndispoGroupe.objects.get(id=pk)
    groupe = indispo.groupe

    if request.method == 'POST':
        indispo.delete()
        return redirect('/iscae_emploi/groupe_indispo/%i' %groupe.id)

    context = {
        'title': 'Delete indispo',
        'item': indispo,
        'groupe':groupe,
    }
    return render(request,'groupe_indispo_delete.html',context)


#ajout cal_exception

@login_required(login_url='/')
def cal_exception_ajout(request,pk):
    cal = CalandrierStandard.objects.get(id=pk)

    form=ExceptioncalandrierForm()
    if request.method == 'POST':
        form = ExceptioncalandrierForm(request.POST)
        if form.is_valid():

            cal_exception = form.save(commit=False)
            cal_exception.date= request.POST['date']
            cal_exception.calandrier=cal
            cal_exception.save()
        return redirect('/iscae_emploi/c_exception/%i' %pk)


    context = {
        'title': 'Ajout cal_exception',
        'cal': cal,
        'form':form
    }
    return render(request,'cal_exception_ajout.html',context)


#edit de cal_exception
@login_required(login_url='/')
def cal_exception_edit(request,pk):

    cal_exception = ExceptionCalandrier.objects.get(id=pk)
    cal = cal_exception.calandrier
    form=ExceptioncalandrierForm(instance=cal_exception)

    if request.method == 'POST':
        form = ExceptioncalandrierForm(request.POST,instance=cal_exception)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/c_exception/%i' %cal.id)
    else:
        form = ExceptioncalandrierForm(instance=cal_exception)

    context = {
        'title': 'Edit cal_exception',
        'form': form,
        'cal_exception':cal_exception,
        'cal':cal
    }
    return render(request,'cal_exception_edit.html',context)


#delete de cal_exception
@login_required(login_url='/')
def cal_exception_delete(request,pk):
    cal_exception = ExceptionCalandrier.objects.get(id=pk)
    cal = cal_exception.calandrier

    if request.method == 'POST':
        cal_exception.delete()
        return redirect('/iscae_emploi/c_exception/%i' %cal.id)

    context = {
        'title': 'Delete cal_exception',
        'item': cal_exception,
        'cal':cal,
    }
    return render(request,'cal_exception_delete.html',context)



#ajout de seance
@login_required(login_url='/')
def seance_ajout(request,pk):

    form=SeanceForm()
    if request.method == 'POST':
        form = SeanceForm(request.POST)
        if form.is_valid():
            seance = form.save(commit=False)
            seance.date= request.POST['seance_date']
            seance.save()
            return redirect('/iscae_emploi/seances/%i' %pk)
    else:
        form = SeanceForm()

    context = {

        'title': 'Ajout seance',
        'form': form,
    }
    return render(request,'seance_ajout.html',context)

#edit de seance
@login_required(login_url='/')
def seance_edit(request,pk):

    seance=Seance.objects.get(id=pk)
    id = seance.cours.groupe.id
    form=SeanceForm(instance=seance)

    if request.method == 'POST':
        form = SeanceForm(request.POST,instance=seance)
        if form.is_valid():
            seanc = form.save(commit=False)
            seanc.date= request.POST['seance_date']
            seanc.save()
            return redirect('/iscae_emploi/seances/%i' %id)
    else:
        form = SeanceForm(instance=seance)

    context = {
        'title': 'Edit seance',
        'form': form,
        'seance':seance,
    }
    return render(request,'seance_edit.html',context)

@login_required(login_url='/')
def seance_delete(request,pk):
    seance = Seance.objects.get(id=pk)

    if request.method == 'POST':
        seance.delete()
        return redirect('/iscae_emploi/seances/%i' %seance.cours.groupe.classe.id)

    context = {
        'title': 'Delete seance',
        'item': seance,
    }
    return render(request,'seance_delete.html',context)


#chevauchement d'une classe
@login_required(login_url='/')
def chevauch(request,pk):
    classe = Classe.objects.get(id=pk)
    ch = GroupeChevauchement.objects.filter(classe_id=pk)


    pagination = Paginator(ch,per_page=5)
    page_number = request.GET.get('page',1)
    page_obj = pagination.get_page(page_number)
    context = {
        'title': 'chevauchement',
        'chevauchement':page_obj.object_list,
        'pagination':pagination,
        'page_number': int(page_number),
        'classe':classe

    }
    return render(request,'chevauch.html',context)


#ajout de chevauchment
@login_required(login_url='/')
def chevauch_ajout(request,pk):
    classe = Classe.objects.get(id=pk)
    grp = Groupe.objects.filter(classe_id=pk)

    if request.method == 'POST':
        form = GroupeChevauchementForm(pk=pk, data=request.POST)
        if form.is_valid():
            ch = form.save(commit=False)
            ch.classe=classe
            ch.save()
            return redirect('/iscae_emploi/chevauch/%i' %pk)
    else:
        form = GroupeChevauchementForm(
            pk=pk,)

    context = {

        'title': 'Ajout chevauchement',
        'form': form,
        'classe':classe
    }
    return render(request,'chevauch_ajout.html',context)




#edit de chevauch
@login_required(login_url='/')
def chevauch_edit(request,pk):

    ch = GroupeChevauchement.objects.get(id=pk)
    id = ch.classe.id
    form=GroupeChevauchementForm(instance=ch,pk=pk)

    if request.method == 'POST':
        form = GroupeChevauchementForm(pk=pk, data=request.POST,instance=ch,)
        if form.is_valid():
            form.save()
            return redirect('/iscae_emploi/chevauch/%i' %id)
    else:
        form = GroupeChevauchementForm(instance=ch,pk=pk)

    context = {
        'title': 'Edit chevauch',
        'form': form,
        'ch':ch,
    }
    return render(request,'chevauch_edit.html',context)

@login_required(login_url='/')
def chevauch_delete(request,pk):
    ch = GroupeChevauchement.objects.get(id=pk)

    if request.method == 'POST':
        ch.delete()
        return redirect('/iscae_emploi/chevauch/%i' %ch.classe.id)

    context = {
        'title': 'Delete chevauch',
        'ch': ch,
    }
    return render(request,'chevauch_delete.html',context)





@login_required(login_url='/')
def seance(request,pk):

    date = datetime.today()
    seance = Seance.objects.filter(cours__groupe__id=pk)

    dd = False


    if request.method == "POST":
        fromdate = request.POST.get('fromdate')

        date = datetime.strptime(fromdate, '%Y-%m-%d')
        nn = date.weekday()
        date_debut = date - timedelta(days=nn+1)
        date_fin = date + timedelta(days=(6-nn))

        seance = Seance.objects.filter(date__range=[date_debut, date_fin])
        if seance.exists():
            dd=True
        else:
            dd=False


        courss = Cours.objects.filter(groupe_id=pk)
        cours = []
        for c in courss:
            cours.append(c)

        groupes = Groupe.objects.filter(id=pk)
        groupe = []
        for c in groupes:
            groupe.append(c)

        matiers = Cours.objects.values_list('matier' ,flat=True)
        matier = []
        g_matiere =[]
        for c in matiers:
            matier.append(c)
            g_matiere.append(Matier.objects.get(id=c))

        profs = Cours.objects.values_list('prof' ,flat=True)
        prof = []
        g_prof= []
        for c in profs:
            prof.append(c)
            g_prof.append(Professeur.objects.get(id=c))

        salles = Cours.objects.values_list('salle' ,flat=True)
        salle = []
        g_salle = []
        for c in salles:
            salle.append(c)
            g_salle.append(Salle.objects.get(id=c))


        #lists Creneaux
        Creneaux=[]
        C = Creneau.objects.all()
        for i in C:
            Creneaux.append(i.nom)


        listt=['ahmed']


        #excel variables #########################################

        #define border formats

        thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                             right=Side(border_style='dashed',color='FF000000'),
                             top=Side(border_style='dashed',color='FF000000'),
                             bottom=Side(border_style='dashed',color='FF000000')
                             )
        thick_border = Border(left=Side(border_style='thick',color='FF000000'),
                             right=Side(border_style='thick',color='FF000000'),
                             top=Side(border_style='thick',color='FF000000'),
                             bottom=Side(border_style='thick',color='FF000000')
                             )
        Double_border = Border(left=Side(border_style='double',color='FF000000'),
                             right=Side(border_style='double',color='FF000000'),
                             top=Side(border_style='double',color='FF000000'),
                             bottom=Side(border_style='double',color='FF000000')
                             )
        fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')


        def column_Number(i):
            return 3 + i//23








    #############excel variables #########################################

        wb = Workbook()
        ws1=wb.active #worksheet
        ws1.title = "Pyxl"

        #dimenssion du cellule
        for i in range(1,3):
            ws1.cell(row=1,column=i).font= Font(size=14,bold=True)


        ws1.column_dimensions['C'].width=40


        #lists des resultats
        listresul=['Emploi']


        for s in seance:
            n=s.créneau.numero
            if(s.date.weekday()==0):
                _ = ws1.cell(column=column_Number(3),row=1+n ,value='%s,%s,%s '%(s.cours, s.cours.prof , s.cours.salle)).border=thin_border
                _ = ws1.cell(column=column_Number(3),row=1+n).fill=fill_cell

            if(s.date.weekday()==1):
                _ = ws1.cell(column=column_Number(3),row=6+n ,value='%s,%s,%s '%(s.cours, s.cours.prof , s.cours.salle)).border=thin_border
                _ = ws1.cell(column=column_Number(3),row=6+n).fill=fill_cell

            if(s.date.weekday()==2):
                _ = ws1.cell(column=column_Number(3),row=11+n ,value='%s,%s,%s '%(s.cours, s.cours.prof , s.cours.salle)).border=thin_border
                _ = ws1.cell(column=column_Number(3),row=11+n).fill=fill_cell

            if(s.date.weekday()==3):
                _ = ws1.cell(column=column_Number(3),row=16+n ,value='%s,%s,%s '%(s.cours, s.cours.prof , s.cours.salle)).border=thin_border
                _ = ws1.cell(column=column_Number(3),row=16+n).fill=fill_cell

            if(s.date.weekday()==4):
                _ = ws1.cell(column=column_Number(3),row=21+n ,value='%s,%s,%s '%(s.cours, s.cours.prof , s.cours.salle)).border=thin_border
                _ = ws1.cell(column=column_Number(3),row=21+n).fill=fill_cell













        ######## Excel ###########

        #define border formats

        thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                             right=Side(border_style='dashed',color='FF000000'),
                             top=Side(border_style='dashed',color='FF000000'),
                             bottom=Side(border_style='dashed',color='FF000000')
                             )
        thick_border = Border(left=Side(border_style='thick',color='FF000000'),
                             right=Side(border_style='thick',color='FF000000'),
                             top=Side(border_style='thick',color='FF000000'),
                             bottom=Side(border_style='thick',color='FF000000')
                             )
        Double_border = Border(left=Side(border_style='double',color='FF000000'),
                             right=Side(border_style='double',color='FF000000'),
                             top=Side(border_style='double',color='FF000000'),
                             bottom=Side(border_style='double',color='FF000000')
                             )




         # ##############   CM ############33

        #les creneaux ....
        for i in range(len(Creneaux)):
            _ = ws1.cell(column=2,row=i+2,value=Creneaux[i]).border=thin_border
        for i in range(len(Creneaux)):
            _ = ws1.cell(column=2,row=i+7,value=Creneaux[i]).border=thin_border
        for i in range(len(Creneaux)):
            _ = ws1.cell(column=2,row=i+12,value=Creneaux[i]).border=thin_border
        for i in range(len(Creneaux)):
            _ = ws1.cell(column=2,row=i+17,value=Creneaux[i]).border=thin_border
        for i in range(len(Creneaux)):
            _ = ws1.cell(column=2,row=i+22,value=Creneaux[i]).border=thin_border


        _ = ws1.cell(column=3,row=1,value='Semaine ' + str(date_debut + timedelta(days=(1)))).border=thick_border
        #les jours
        ws1.merge_cells('A2:A6')
        ws1['A2']='Lundi'
        ws1.merge_cells('A7:A11')
        ws1['A7']='mardi'
        ws1.merge_cells('A12:A16')
        ws1['A12']='Mercredi'
        ws1.merge_cells('A17:A21')
        ws1['A17']='Jeudi'
        ws1.merge_cells('A22:A26')
        ws1['A22']='Vendredi'
        for i in range(2,27):
            _ = ws1.cell(column=1,row=i).border=thick_border
        ws1['A1']='Jours'
        ws1['B1']='Heurs'

        wb.save('groupe_semaine_Emploi.csv')


        return render(request,'seance.html',{'seance':seance,
                                             'groupe':pk,
                                             'dd':dd})
    else:
        pagination = Paginator(seance,per_page=5)
        page_number = request.GET.get('page',1)
        page_obj = pagination.get_page(page_number)


        context = {
            'groupe':pk,
            'title': 'seance',
            'seance': page_obj.object_list,
            'pagination':pagination,
            'page_number': int(page_number),
            'dd':dd
        }
        return render(request,'seance.html',context)






@login_required(login_url='/')
def seance_download(request):
    # fill these variables with real values
    fl_path = 'D:/ahmed PFE/emploi/emploi/groupe_semaine_Emploi.csv'
    filename = 'groupe_semaine_Emploi.csv'

    fl = open(fl_path,'r',errors='ignore')

    mime_type, _ = mimetypes.guess_type(fl_path)
    response = HttpResponse(fl, content_type=mime_type)
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    return response









@login_required(login_url='/')
def emploi(request,pk):
    classe = Classe.objects.get(id = pk)

    nombre_creneaux = 0
    date_valide = []


    def n_creneaux(li):
        jour_numero = [0,1,2,3]
        n=0
        for date in li :
            if date.weekday() in jour_numero:
                date_valide.append(date)
                n = n + 5
            elif date.weekday()==4:
                date_valide.append(date)
                n = n + 2
        return  n

    context={}

    def generation(n_creneaux,d_debut,d_fin,pk):

        classe = Classe.objects.filter(id=pk)
        print('**************************************************************************************')

        groupes = Cours.objects.filter(groupe__classe_id=pk).values_list('groupe' ,flat=True)
        print(groupes)
        groupe = []
        groupeId = []
        groupe_indispo=[]
        for c in groupes:
            if c not in groupeId:
                groupe.append(Groupe.objects.get(id=c))
                groupeId.append(c)
                print(c)
        for p in groupeId:
            g = IndispoGroupe.objects.filter(groupe_id = p)
            groupe_indispo.append(g)
            for i in g:
                print(i.groupe , i.date)
                for j in i.créneaux.all():
                    print(j.numero)

        courss = Cours.objects.filter(groupe_id__in=groupeId )
        cours = []
        for c in courss:
            cours.append(c)
            print(c)

        matiers = Cours.objects.filter(groupe_id__in=groupeId).values_list('matier' ,flat=True)
        matier = []
        g_matiere =[]
        for c in matiers:
            matier.append(c)
            g_matiere.append(Matier.objects.get(id=c))
            print(c)

        profs = Cours.objects.filter(groupe_id__in=groupeId).values_list('prof' ,flat=True)
        prof =[]
        g_prof= []
        g_prof_indispo=[]
        for c in profs:
            prof.append(c)
            print(c)
            g_prof.append(Professeur.objects.get(id=c))
        for p in g_prof:
            g = IndispoProf.objects.filter(prof_id = p)
            g_prof_indispo.append(g)
            for i in g:
                print(i.prof , i.date)
                for j in i.créneaux.all():
                    print(j.numero)



        salles = Cours.objects.filter(groupe_id__in=groupeId).values_list('salle' ,flat=True)
        salle =[]
        g_salle = []
        for c in salles:
            salle.append(c)
            g_salle.append(Salle.objects.get(id=c))


        #lists Creneaux
        Creneaux=[]
        C = Creneau.objects.all()
        for i in C:
            Creneaux.append(i.nom)

        chevauchement_list = GroupeChevauchement.objects.filter(classe_id=pk)



        listt=['ahmed']
        context ={
            'profs' : profs ,
            'matier' : g_matiere ,
            'groupe' : groupes,
            'cours' :courss,
            'salle':salles,
            'classe':classe,

        }

        #excel variables #########################################

        #define border formats

        thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                             right=Side(border_style='dashed',color='FF000000'),
                             top=Side(border_style='dashed',color='FF000000'),
                             bottom=Side(border_style='dashed',color='FF000000')
                             )
        thick_border = Border(left=Side(border_style='thick',color='FF000000'),
                             right=Side(border_style='thick',color='FF000000'),
                             top=Side(border_style='thick',color='FF000000'),
                             bottom=Side(border_style='thick',color='FF000000')
                             )
        Double_border = Border(left=Side(border_style='double',color='FF000000'),
                             right=Side(border_style='double',color='FF000000'),
                             top=Side(border_style='double',color='FF000000'),
                             bottom=Side(border_style='double',color='FF000000')
                             )
        fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')


        def column_Number(i):
            return 3 + i//22


        # Create the linear solver with the GLOP backend.
        solver = pywraplp.Solver('cas1',
                                 pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)

        # Create the variables x and y.


        h = n_creneaux # nombre creneaux

        m = matiers.count()   # nombre  matieres
        n = groupe.__len__()   # nombre groupes
        p = profs.count()  # nombre enseignants
        s=np.zeros(((n,m,p)))
        print(m,n,p)

        print('****************'+str(p))

        #les charge de matier
        Ch = []
        for c in courss:
            Ch.append(c.charge)




        # nombre de courss par groupe/mat/ens s[n][m][p]
        for i in range(0,n):
            for j in range(0,m):
                for u in range(0,p):
                    if j==u:
                        s[i][j][u]=Ch[j]




    #############excel variables #########################################


        wb = Workbook()
        ww = []
        t = wb.active
        print(" groupe.index(0).nom ***********"  + groupe[0].nom)
        t.title = groupe[0].nom
        ww.append(t)

        for e in range(1,n):
            w = wb.create_sheet()
            w.title = groupe[e].nom
            ww.append(w)
        print(ww)

  # ws =wb.active #worksheet
 # ws.title = "Pyxl"


        for e in ww:
            #dimenssion du cellule
            for i in range(1,3):
                e.cell(row=1,column=i).font= Font(size=14,bold=True)

            for u in ['c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r']:
                 e.column_dimensions[u].width=40


        #lists des resultats
        listresul=['Emploi']

        #les semaines*************************8

        Semaine=['Rien']
        nombre_semaine = 0
        weeks = rrule.rrule(rrule.WEEKLY, dtstart=d_debut, until=d_fin)
        nombre_semaine=weeks.count()


        debut = date_debut - timedelta(days=d_debut.weekday())

        for i in range(1,nombre_semaine+1):
            print(i)
            if i==1 :
                Semaine.append('Semaine du ' + str(debut))

            else:
                Semaine.append('Semaine du ' + str(debut + timedelta(days=7*(i-1))))

        #disponibiliter de l'enseignant *****************
        jour_numero = [0,1,2,3]

        def nn_creneaux(li):

            b=0
            for date in li :
                if date.weekday() in jour_numero:
                    date_valide.append(date)
                    b = b + 5
                elif date.weekday()==4:
                    date_valide.append(date)
                    b = b + 2
            return  b

        d = np.zeros((p,h))
        print(p,h)
        for i in range(0,p):
            for j in range(0,h):
                d[i,j]=1
        print(d)




        for i in g_prof_indispo:
            print(i)
            for j in i:
                if j.date in date_valide:
                    l = pd.date_range(date_debut, j.date, freq='D')
                    b = nn_creneaux(l)

                    if j.date.weekday() in jour_numero:
                        for y in j.créneaux.all():
                            d[g_prof_indispo.index(i),b-(5-y.numero)-1]=0
                    elif j.date.weekday() == 4:
                        for y in j.créneaux.all():
                            d[g_prof_indispo.index(i),b-(2-y.numero)-1]=0

        print(d)
        print('***********8')




        #Creneaux disponible pour pes(IRT)
        c0=[0,1,2,3,5,15,16,17,18]

        for i in range(0,9):

            c0.append(c0[i]+23)
            c0.append(c0[i]+46)
            c0.append(c0[i]+69)
            c0.append(c0[i]+92)
            c0.append(c0[i]+115)
            c0.append(c0[i]+138)
            c0.append(c0[i]+161)
            c0.append(c0[i]+184)
            c0.append(c0[i]+207)
            c0.append(c0[i]+230)
            c0.append(c0[i]+253)
            c0.append(c0[i]+276)
            c0.append(c0[i]+299)
            c0.append(c0[i]+322)
            c0.append(c0[i]+345)






        ########### Disponibilite du groupe ########################

        dg =np.zeros((2,h))
        #dg[0,0]=1
        for j in range(0,h):
            dg[0,j]=1
            dg[1,j]=1
        print(dg)




        ######################### exception################3

        exceptt = ExceptionCalandrier.objects.all()
        exception = []
        for i in exceptt:
            exception.append(i)



        print(exception)
        print('exception*********************88')

        for j in exception:
            print(j.date,j.créneau.all())
            if j.date in date_valide:
                    l = pd.date_range(date_debut, j.date, freq='D')
                    b = nn_creneaux(l)
                    for k in groupe_indispo:
                        if j.date.weekday() in jour_numero:
                            print(j.créneau.all())
                            for y in j.créneau.all():
                                print('numero du creneaux'+str(y.numero))
                                print('index'+ str(groupe_indispo.index(k)))
                                dg[groupe_indispo.index(k),b-(5-y.numero)-1]=0
                        elif j.date.weekday() == 4 :
                            for y in j.créneau.all():
                                if y.numero < 3:
                                    print(groupe_indispo.index(k),b,y.numero)
                                    dg[groupe_indispo.index(i),b-(2-y.numero)-1]=0
        print(dg)


        for i in groupe_indispo:
            for j in i:
                if j.date in date_valide:
                    l = pd.date_range(date_debut, j.date, freq='D')
                    b = nn_creneaux(l)

                    print(j.date ,j.date.weekday(), b)
                    print('index'+ str(groupe_indispo.index(i)))

                    if j.date.weekday() in jour_numero:
                        print(j.créneaux.all())
                        for y in j.créneaux.all():

                            print('numero du creneaux'+str(y.numero))
                            dg[groupe_indispo.index(i),b-(5-y.numero)-1]=0
                    elif j.date.weekday() == 4 :
                        for y in j.créneaux.all():
                            if y.numero < 3:
                                print(groupe_indispo.index(i),b,y.numero)
                                dg[groupe_indispo.index(i),b-(2-y.numero)-1]=0
        print(dg)




        #chevauchment
        v =np.zeros((n,n))
        #dg[0,0]=1
        for j in range(0,n):
            for i in range(0,n):
                v[j,i]=1

        if n!=1 :
            for u in chevauchement_list:
                v[u.groupe1.numero-1,u.groupe2.numero-1]=0
        print("***********************777777777777777**********************************************************************")
        print(v)
        x = {}

        for i in range(1,n+1) :
            for j in  range(1,m+1):
                for l in  range(1,p+1):
                    for k in  range(1,h+1):
                        x[(i, j,
                            l,k)] = solver.IntVar(0, 1, 'x_i%ij%il%ik%i' % (i, j, l,k));

        print('Number of variables =', solver.NumVariables())

        #meme professeur 2 fois pour le meme creneau est interdit
        print(p,h,n,m)
        for l in range(1,p+1):
            for k in  range(1,h+1):
                solver.Add(sum(x[(i, j, l, k)] for i in range(1,n+1) for j in range(1,m+1) ) <= 1)

        #meme groupe 2 fois pour le meme creneau est interdit
        for i in range(1,n+1):
            for k in  range(1,h+1):
                solver.Add(sum(x[(i, j, l, k)] for l in range(1,p+1) for j in range(1,m+1) ) <= 1)


        # (3) toutes les courss doivent etre dispenses
        for i in range(1,n+1):
            for j in  range(1,m+1):
                for l in  range(1,p+1):
                    solver.Add(sum(x[(i, j, l, k)] for k in range(1,h+1) ) == s[i-1][j-1][l-1])

        # (4) Disponibilite de l'enseignant
        for l in range(1,p+1):
            for k in  range(1,h+1):
                solver.Add(sum(x[(i, j, l, k)] for i in range(1,n+1) for j in range(1,m+1) ) <= d[l-1][k-1])

        # (5) Disponibilite du groupe
        print(m,h,p,n)
        for i in range(1,n+1):
            for k in  range(1,h+1):
                solver.Add(sum(x[(i, j, l, k)] for j in range(1,m+1) for l in range(1,p+1) ) <= dg[i-1][k-1])

        # (6) chevauchement  entre les groupes

        for i1 in range(1,n+1):
            for i2 in  range(1,n+1):
                for k in  range(1,h+1):
                    if(i1 < i2 and v[i1-1][i2-1]==1):
                        g1 = sum(x[(i1, j, l, k)] for j in range(1,m+1) for l in range(1,p+1) )
                        g2 = sum(x[(i2, j, l, k)] for j in range(1,m+1) for l in range(1,p+1))
                        solver.Add(g1 + g2  <= 1)

        solver.Solve()



        for i in range(1,n+1) :
            for j in  range(1,m+1):
                for l in  range(1,p+1):
                    for k in  range(1,h+1):
                        if(x[i,j,l,k].solution_value() == 1):
                            print('x[%s,%s,%s,%i] ='% (groupe[i-1], matier[j-1], prof[l-1],k), x[i,j,l,k].solution_value())
                            listt.append(str(groupe[i-1]) + " " + str(matier[j-1]) + " " + str(prof[l-1]) ) #+" "+ matier[j-1]+" "+ professeurs[l-1])
                            if k in range(0,22):
                                _ = ww[i-1].cell(column=column_Number(k),row=k+1,value='%s,%s,%s '%(groupe[i-1], (g_matiere[j-1]).nom, (g_prof[l-1]).nom)).border=thin_border
                                _ = ww[i-1].cell(column=column_Number(k),row=k+1).fill=fill_cell
                            elif k in range(22,h):
                                print(int(k)%22)
                                _ = ww[i-1].cell(column=column_Number(k),
                                             row=int((int(k)%22)+2),value='%s,%s,%s'%(groupe[i-1], (g_matiere[j-1]).nom, (g_prof[l-1]).nom)).border=thin_border
                                _ = ww[i-1].cell(column=column_Number(k),row=k%22+2).fill=fill_cell



        ######## Excel ###########

        #define border formats

        thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                             right=Side(border_style='dashed',color='FF000000'),
                             top=Side(border_style='dashed',color='FF000000'),
                             bottom=Side(border_style='dashed',color='FF000000')
                             )
        thick_border = Border(left=Side(border_style='thick',color='FF000000'),
                             right=Side(border_style='thick',color='FF000000'),
                             top=Side(border_style='thick',color='FF000000'),
                             bottom=Side(border_style='thick',color='FF000000')
                             )
        Double_border = Border(left=Side(border_style='double',color='FF000000'),
                             right=Side(border_style='double',color='FF000000'),
                             top=Side(border_style='double',color='FF000000'),
                             bottom=Side(border_style='double',color='FF000000')
                             )




         # ##############   CM ############33
        for e in ww:
            #les creneaux ....
            for i in range(len(Creneaux)):
                _ = e.cell(column=2,row=i+2,value=Creneaux[i]).border=thin_border
            for i in range(len(Creneaux)):
                _ = e.cell(column=2,row=i+7,value=Creneaux[i]).border=thin_border
            for i in range(len(Creneaux)):
                _ = e.cell(column=2,row=i+12,value=Creneaux[i]).border=thin_border
            for i in range(len(Creneaux)):
                _ = e.cell(column=2,row=i+17,value=Creneaux[i]).border=thin_border
            for i in range(len(Creneaux)):
                _ = e.cell(column=2,row=i+22,value=Creneaux[i]).border=thin_border


        for f in ww:
            #le ligne des semaines
            for row in range(1,len(Semaine)):
                _ = f.cell(column=row+2,row=1,value=Semaine[row]).border=thick_border
            #les jours
            f.merge_cells('A2:A6')
            f['A2']='Lundi'
            f.merge_cells('A7:A11')
            f['A7']='mardi'
            f.merge_cells('A12:A16')
            f['A12']='Mercredi'
            f.merge_cells('A17:A21')
            f['A17']='Jeudi'
            f.merge_cells('A22:A26')
            f['A22']='Vendredi'
            for i in range(2,27):
                _ = f.cell(column=1,row=i).border=thick_border
            f['A1']='Jours'
            f['B1']='Heurs'

            wb.save('%s.xlsx' %e.title)


    if request.method == 'POST':
        d = request.POST['date_debut']
        date_debut = datetime.strptime(d, '%Y-%m-%d')
        f = request.POST['date_fin']
        date_fin = datetime.strptime(f, '%Y-%m-%d')
        print(date_fin, date_debut)
        li = pd.date_range(date_debut, date_fin, freq='D')
        nombre_creneaux = n_creneaux(li)
        generation(nombre_creneaux,date_debut,date_fin,pk)
    print(nombre_creneaux)



    return render(request,'emploi.html',context)



@login_required(login_url='/')
def emploi_generation(request,pk):

    courss = Cours.objects.filter(groupe_id=pk)
    cours = []
    for c in courss:
        cours.append(c)

    groupes = Groupe.objects.filter(id=pk)
    groupe = []
    for c in groupes:
        groupe.append(c)

    matiers = Cours.objects.values_list('matier' ,flat=True)
    matier = []
    g_matiere =[]
    for c in matiers:
        matier.append(c)
        g_matiere.append(Matier.objects.get(id=c))

    profs = Cours.objects.values_list('prof' ,flat=True)
    prof = []
    g_prof= []

    for c in profs:
        prof.append(c)
        g_prof.append(Professeur.objects.get(id=c))


    salles = Cours.objects.values_list('salle' ,flat=True)
    salle = []
    g_salle = []
    for c in salles:
        salle.append(c)
        g_salle.append(Salle.objects.get(id=c))


    #lists Creneaux
    Creneaux=[]
    C = Creneau.objects.all()
    for i in C:
        Creneaux.append(i.nom)


    listt=['ahmed']
    context ={
        'profs' : profs ,
        'matier' : g_matiere ,
        'groupe' : groupes,
        'cours' :courss,
        'salle':salles,

    }

    #excel variables #########################################

    #define border formats

    thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                         right=Side(border_style='dashed',color='FF000000'),
                         top=Side(border_style='dashed',color='FF000000'),
                         bottom=Side(border_style='dashed',color='FF000000')
                         )
    thick_border = Border(left=Side(border_style='thick',color='FF000000'),
                         right=Side(border_style='thick',color='FF000000'),
                         top=Side(border_style='thick',color='FF000000'),
                         bottom=Side(border_style='thick',color='FF000000')
                         )
    Double_border = Border(left=Side(border_style='double',color='FF000000'),
                         right=Side(border_style='double',color='FF000000'),
                         top=Side(border_style='double',color='FF000000'),
                         bottom=Side(border_style='double',color='FF000000')
                         )
    fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')


    def column_Number(i):
        return 3 + i//23


    # Create the linear solver with the GLOP backend.
    solver = pywraplp.Solver('cas1',
                             pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)

    # Create the variables x and y.


    h = 368 # nombre creneaux
    m = matiers.count()   # nombre  matieres
    n = groupes.count()   # nombre groupes
    p = profs.count()  # nombre enseignants
    s=np.zeros(((n,m,p)))
    print(m,n,p)

    #les charge de matier
    Ch = []
    for c in courss:
        Ch.append(c.charge)




    # nombre de courss par groupe/mat/ens s[n][m][p]
    for i in range(0,n):
        for j in range(0,m):
            for u in range(0,p):
                if j==u:
                    s[i][j][u]=Ch[j]
    print(s)



#############excel variables #########################################

    wb = Workbook()
    ws1=wb.active #worksheet
    ws1.title = "Pyxl"

    #dimenssion du cellule
    for i in range(1,3):
        ws1.cell(row=1,column=i).font= Font(size=14,bold=True)

    for u in ['c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r']:
         ws1.column_dimensions[u].width=40


    #lists des resultats
    listresul=['Emploi']

    #les semaines
    Semaine=['S1','S2','S3','S4','S5','S6','S7','S8','S9','S10','S11','S12','S13','S14','S15','S16']


    #disponibiliter de l'enseignant
    d = np.zeros((7,h))
    for i in range(0,7):
        for j in range(0,h):
            d[i,j]=1
    print(d[1].__len__())

    #Creneaux disponible pour pes(IRT)
    c0=[0,1,2,3,5,15,16,17,18]

    for i in range(0,9):

        c0.append(c0[i]+23)
        c0.append(c0[i]+46)
        c0.append(c0[i]+69)
        c0.append(c0[i]+92)
        c0.append(c0[i]+115)
        c0.append(c0[i]+138)
        c0.append(c0[i]+161)
        c0.append(c0[i]+184)
        c0.append(c0[i]+207)
        c0.append(c0[i]+230)
        c0.append(c0[i]+253)
        c0.append(c0[i]+276)
        c0.append(c0[i]+299)
        c0.append(c0[i]+322)
        c0.append(c0[i]+345)
    print(c0)




    ########### Disponibilite du groupe ########################

    dg =np.zeros((1,h))
    #dg[0,0]=1
    for j in range(0,h):
        if j in c0:
            dg[0,j]=1

    print(dg)

    #chevauchment
    v = [[1,1,1,1,1]]

    x = {}

    for i in range(1,n+1) :
        for j in  range(1,m+1):
            for l in  range(1,p+1):
                for k in  range(1,h+1):
                    x[(i, j,
                        l,k)] = solver.IntVar(0, 1, 'x_i%ij%il%ik%i' % (i, j, l,k));

    print('Number of variables =', solver.NumVariables())

    #meme professeur 2 fois pour le meme creneau est interdit
    for l in range(1,p+1):
        for k in  range(1,h+1):
            solver.Add(sum(x[(i, j, l, k)] for i in range(1,n+1) for j in range(1,m+1) ) <= 1)

    #meme groupe 2 fois pour le meme creneau est interdit
    for i in range(1,n+1):
        for k in  range(1,h+1):
            solver.Add(sum(x[(i, j, l, k)] for l in range(1,p+1) for j in range(1,m+1) ) <= 1)


    # (3) toutes les courss doivent etre dispenses
    for i in range(1,n+1):
        for j in  range(1,m+1):
            for l in  range(1,p+1):
                solver.Add(sum(x[(i, j, l, k)] for k in range(1,h+1) ) == s[i-1][j-1][l-1])

    # (4) Disponibilite de l'enseignant
    #for l in range(1,p+1):
     #   for k in  range(1,h+1):
      #      solver.Add(sum(x[(i, j, l, k)] for i in range(1,n+1) for j in range(1,m+1) ) <= d[l-1][k-1])
    # (5) Disponibilite du groupe
    for i in range(1,n+1):
        for k in  range(1,h+1):
            solver.Add(sum(x[(i, j, l, k)] for j in range(1,m+1) for l in range(1,p+1) ) <= dg[i-1][k-1])
    # (6) chevauchement  entre les groupes
    for i1 in range(1,n+1):
        for i2 in  range(1,n+1):
            for k in  range(1,h+1):
                if(i1 < i2 and v[i1-1][i2-1]==1):
                    g1 = sum(x[(i1, j, l, k)] for j in range(1,m+1) for l in range(1,p+1) )
                    g2 = sum(x[(i2, j, l, k)] for j in range(1,m+1) for l in range(1,p+1))
                    solver.Add(g1 + g2  <= 1)

    solver.Solve()

    print('Solution:')

    for i in range(1,n+1) :
        for j in  range(1,m+1):
            for l in  range(1,p+1):
                for k in  range(1,h+1):
                    if(x[i,j,l,k].solution_value() == 1):
                        print('x[%s,%s,%s,%i] ='% (groupe[i-1], matier[j-1], prof[l-1],k), x[i,j,l,k].solution_value())
                        listt.append(str(groupe[i-1]) + " " + str(matier[j-1]) + " " + str(prof[l-1]) ) #+" "+ matier[j-1]+" "+ professeurs[l-1])
                        if k in range(0,23):
                            _ = ws1.cell(column=column_Number(k),row=k+1,value='%s,%s,%s '%(groupe[i-1], (g_matiere[j-1]).nom, (g_prof[l-1]).nom)).border=thin_border
                            _ = ws1.cell(column=column_Number(k),row=k+1).fill=fill_cell
                        elif k in range(23,368):
                            print(int(k)%23)
                            _ = ws1.cell(column=column_Number(k),
                                         row=int(int(k)%23+1),value='%s,%s,%s'%(groupe[i-1], (g_matiere[j-1]).nom, (g_prof[l-1]).nom)).border=thin_border
                            _ = ws1.cell(column=column_Number(k),row=int(int(k)%23)+1).fill=fill_cell



    ######## Excel ###########

    #define border formats

    thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                         right=Side(border_style='dashed',color='FF000000'),
                         top=Side(border_style='dashed',color='FF000000'),
                         bottom=Side(border_style='dashed',color='FF000000')
                         )
    thick_border = Border(left=Side(border_style='thick',color='FF000000'),
                         right=Side(border_style='thick',color='FF000000'),
                         top=Side(border_style='thick',color='FF000000'),
                         bottom=Side(border_style='thick',color='FF000000')
                         )
    Double_border = Border(left=Side(border_style='double',color='FF000000'),
                         right=Side(border_style='double',color='FF000000'),
                         top=Side(border_style='double',color='FF000000'),
                         bottom=Side(border_style='double',color='FF000000')
                         )




     # ##############   CM ############33

    #les creneaux ....
    for i in range(len(Creneaux)):
        _ = ws1.cell(column=2,row=i+2,value=Creneaux[i]).border=thin_border
    for i in range(len(Creneaux)):
        _ = ws1.cell(column=2,row=i+7,value=Creneaux[i]).border=thin_border
    for i in range(len(Creneaux)):
        _ = ws1.cell(column=2,row=i+12,value=Creneaux[i]).border=thin_border
    for i in range(len(Creneaux)):
        _ = ws1.cell(column=2,row=i+17,value=Creneaux[i]).border=thin_border
    for i in range(len(Creneaux)):
        _ = ws1.cell(column=2,row=i+22,value=Creneaux[i]).border=thin_border

    #le ligne des semaines
    for row in range(len(Semaine)):
        _ = ws1.cell(column=row+3,row=1,value=Semaine[row]).border=thick_border
    #les jours
    ws1.merge_cells('A2:A6')
    ws1['A2']='Lundi'
    ws1.merge_cells('A7:A11')
    ws1['A7']='mardi'
    ws1.merge_cells('A12:A16')
    ws1['A12']='Mercredi'
    ws1.merge_cells('A17:A21')
    ws1['A17']='Jeudi'
    ws1.merge_cells('A22:A26')
    ws1['A22']='Vendredi'
    for i in range(2,27):
        _ = ws1.cell(column=1,row=i).border=thick_border
    ws1['A1']='Jours'
    ws1['B1']='Heurs'

    wb.save('pes_Emploi.xlsx')

    return render(request,'viewer.html',context)

